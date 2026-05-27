"""
core/exporter.py
ParsedEntry リストを Excel (.xlsx) に出力するモジュール。

列順・シート名は clients.yaml から読み込む。
全て純粋関数。副作用なし (ファイル書き込みを除く)。
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
import yaml

from core.parser import ParsedEntry

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

# ParsedEntry フィールド名 → Excel列名 のマッピング
_FIELD_TO_COLUMN: dict[str, str] = {
    "receipt_number": "受付番号",
    "receipt_date":   "受付日",
    "category":       "区分",
    "purpose":        "目的",
    "property_type":  "種別",
    "location":       "所在",
    "remarks":        "備考",
}

# Excel列名 → ParsedEntry フィールド名 (逆引き)
_COLUMN_TO_FIELD: dict[str, str] = {v: k for k, v in _FIELD_TO_COLUMN.items()}

# デフォルト列順 (clients.yaml がない場合のフォールバック)
_DEFAULT_COLUMN_ORDER: list[str] = [
    "受付番号", "受付日", "区分", "目的", "種別", "所在", "備考",
]

# ---------------------------------------------------------------------------
# DataFrame 変換
# ---------------------------------------------------------------------------


def to_dataframe(
    entries: list[ParsedEntry],
    column_order: Optional[list[str]] = None,
) -> pd.DataFrame:
    """ParsedEntry のリストを pandas DataFrame に変換する。

    Args:
        entries: 変換対象のエントリリスト。
        column_order: 出力する列名の順序リスト。
                      None の場合は _DEFAULT_COLUMN_ORDER を使用。

    Returns:
        指定列順の DataFrame。
    """
    cols = column_order or _DEFAULT_COLUMN_ORDER

    rows = []
    for e in entries:
        row = {}
        for col in cols:
            field = _COLUMN_TO_FIELD.get(col)
            row[col] = getattr(e, field, "") if field else ""
        rows.append(row)

    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
# Excel 書き込み
# ---------------------------------------------------------------------------


def _apply_styles(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """ヘッダー行のスタイルを設定する (太字・背景色・列幅)。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")   # 紺
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(size=10)
    ALIGN_LEFT  = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    # 列幅プリセット (列名 → 文字数)
    COL_WIDTHS: dict[str, int] = {
        "受付番号": 14,
        "受付日":   10,
        "区分":      8,
        "目的":     28,
        "種別":      8,
        "所在":     36,
        "備考":     10,
    }

    # ヘッダー行
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_CENTER

    # データ行
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = ALIGN_LEFT

    # 列幅
    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        width = COL_WIDTHS.get(str(cell.value), 14)
        ws.column_dimensions[col_letter].width = width

    # 行高さ (ヘッダー)
    ws.row_dimensions[1].height = 18

    # ウィンドウ枠の固定 (ヘッダー行)
    ws.freeze_panes = "A2"

    # オートフィルタ
    ws.auto_filter.ref = ws.dimensions


def to_excel_bytes(
    entries: list[ParsedEntry],
    sheet_name: str = "相続案件リスト",
    column_order: Optional[list[str]] = None,
) -> bytes:
    """ParsedEntry リストを Excel ファイルのバイト列として返す。

    Streamlit の st.download_button に直接渡せる形式。

    Args:
        entries: 出力対象のエントリリスト。
        sheet_name: Excelシート名。
        column_order: 列順リスト。None の場合はデフォルト列順。

    Returns:
        xlsx 形式のバイト列。
    """
    df = to_dataframe(entries, column_order=column_order)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ヘッダー行
    ws.append(df.columns.tolist())

    # データ行
    for _, row in df.iterrows():
        ws.append(row.tolist())

    _apply_styles(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(
    entries: list[ParsedEntry],
    output_path: Path,
    sheet_name: str = "相続案件リスト",
    column_order: Optional[list[str]] = None,
) -> None:
    """ParsedEntry リストを Excel ファイルとして保存する。

    Args:
        entries: 出力対象のエントリリスト。
        output_path: 保存先パス (.xlsx)。
        sheet_name: Excelシート名。
        column_order: 列順リスト。None の場合はデフォルト列順。
    """
    output_path.write_bytes(
        to_excel_bytes(entries, sheet_name=sheet_name, column_order=column_order)
    )


# ---------------------------------------------------------------------------
# clients.yaml ユーティリティ
# ---------------------------------------------------------------------------


def load_excel_config(
    yaml_path: Path,
    client_id: str = "harada_tatemono",
) -> dict[str, object]:
    """clients.yaml から Excel 出力設定を読み込む。

    Args:
        yaml_path: clients.yaml のパス。
        client_id: YAMLのクライアントキー。

    Returns:
        {"sheet_name": str, "column_order": list[str]} の辞書。
    """
    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    excel_cfg = data[client_id].get("excel", {})
    return {
        "sheet_name":   excel_cfg.get("sheet_name", "相続案件リスト"),
        "column_order": excel_cfg.get("column_order", _DEFAULT_COLUMN_ORDER),
    }
