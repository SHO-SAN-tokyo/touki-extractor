"""
tests/test_exporter.py
core/exporter.py のユニットテスト。
"""

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from core.exporter import (
    _DEFAULT_COLUMN_ORDER,
    load_excel_config,
    save_excel,
    to_dataframe,
    to_excel_bytes,
)
from core.parser import ParsedEntry


# ---------------------------------------------------------------------------
# テスト用フィクスチャ
# ---------------------------------------------------------------------------

def _e(
    receipt_number="第10104号",
    receipt_date="3月2日",
    category="単独",
    purpose="所有権移転・相続",
    property_type="土地",
    location="堺市堺区榎元町一丁10",
    remarks="",
) -> ParsedEntry:
    return ParsedEntry(
        receipt_number=receipt_number,
        receipt_date=receipt_date,
        category=category,
        purpose=purpose,
        property_type=property_type,
        location=location,
        remarks=remarks,
    )


_SAMPLE_ENTRIES = [
    _e(receipt_number="第10104号", location="堺市堺区榎元町一丁10"),
    _e(receipt_number="第10114号", location="堺市北区金岡町1100-29", remarks="外1"),
    _e(receipt_number="第10124号", category="連先",
       location="堺市西区浜寺諏訪森町西四丁306-5", remarks="外1"),
]


# ---------------------------------------------------------------------------
# to_dataframe
# ---------------------------------------------------------------------------


class TestToDataframe:
    def test_row_count(self):
        df = to_dataframe(_SAMPLE_ENTRIES)
        assert len(df) == 3

    def test_default_columns(self):
        df = to_dataframe(_SAMPLE_ENTRIES)
        assert list(df.columns) == _DEFAULT_COLUMN_ORDER

    def test_custom_column_order(self):
        order = ["受付番号", "所在", "目的"]
        df = to_dataframe(_SAMPLE_ENTRIES, column_order=order)
        assert list(df.columns) == order

    def test_cell_values(self):
        df = to_dataframe(_SAMPLE_ENTRIES)
        assert df.iloc[0]["受付番号"] == "第10104号"
        assert df.iloc[0]["受付日"]   == "3月2日"
        assert df.iloc[0]["区分"]     == "単独"
        assert df.iloc[0]["目的"]     == "所有権移転・相続"
        assert df.iloc[0]["種別"]     == "土地"
        assert df.iloc[0]["所在"]     == "堺市堺区榎元町一丁10"
        assert df.iloc[0]["備考"]     == ""

    def test_remarks_preserved(self):
        df = to_dataframe(_SAMPLE_ENTRIES)
        assert df.iloc[1]["備考"] == "外1"

    def test_empty_entries(self):
        df = to_dataframe([])
        assert len(df) == 0
        assert list(df.columns) == _DEFAULT_COLUMN_ORDER

    def test_order_preserved(self):
        df = to_dataframe(_SAMPLE_ENTRIES)
        assert list(df["受付番号"]) == ["第10104号", "第10114号", "第10124号"]


# ---------------------------------------------------------------------------
# to_excel_bytes
# ---------------------------------------------------------------------------


class TestToExcelBytes:
    def _load_wb(self, entries, **kwargs):
        data = to_excel_bytes(entries, **kwargs)
        return openpyxl.load_workbook(io.BytesIO(data))

    def test_returns_bytes(self):
        result = to_excel_bytes(_SAMPLE_ENTRIES)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self):
        data = to_excel_bytes(_SAMPLE_ENTRIES)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_sheet_name_default(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        assert "相続案件リスト" in wb.sheetnames

    def test_sheet_name_custom(self):
        wb = self._load_wb(_SAMPLE_ENTRIES, sheet_name="テスト")
        assert "テスト" in wb.sheetnames

    def test_header_row(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, len(_DEFAULT_COLUMN_ORDER) + 1)]
        assert headers == _DEFAULT_COLUMN_ORDER

    def test_data_row_count(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        # ヘッダー1行 + データ3行 = 4行
        assert ws.max_row == 4

    def test_data_cell_values(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        # 2行目 (最初のデータ行)
        assert ws.cell(2, 1).value == "第10104号"
        assert ws.cell(2, 6).value == "堺市堺区榎元町一丁10"

    def test_remarks_cell(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        # 3行目 (第10114号, 備考=外1)
        assert ws.cell(3, 7).value == "外1"

    def test_empty_remarks_is_empty_string(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        assert ws.cell(2, 7).value in ("", None)

    def test_header_is_bold(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True

    def test_freeze_panes_set(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_autofilter_set(self):
        wb = self._load_wb(_SAMPLE_ENTRIES)
        ws = wb.active
        assert ws.auto_filter.ref is not None

    def test_empty_entries(self):
        data = to_excel_bytes([])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        # ヘッダーのみ
        assert ws.max_row == 1

    def test_custom_column_order(self):
        order = ["受付番号", "所在", "区分"]
        wb = self._load_wb(_SAMPLE_ENTRIES, column_order=order)
        ws = wb.active
        assert ws.cell(1, 1).value == "受付番号"
        assert ws.cell(1, 2).value == "所在"
        assert ws.cell(1, 3).value == "区分"


# ---------------------------------------------------------------------------
# save_excel
# ---------------------------------------------------------------------------


class TestSaveExcel:
    def test_creates_file(self, tmp_path):
        out = tmp_path / "test_output.xlsx"
        save_excel(_SAMPLE_ENTRIES, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_file_is_valid_xlsx(self, tmp_path):
        out = tmp_path / "test_output.xlsx"
        save_excel(_SAMPLE_ENTRIES, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) == 1

    def test_sheet_name_applied(self, tmp_path):
        out = tmp_path / "out.xlsx"
        save_excel(_SAMPLE_ENTRIES, out, sheet_name="出力テスト")
        wb = openpyxl.load_workbook(out)
        assert "出力テスト" in wb.sheetnames

    def test_data_integrity(self, tmp_path):
        out = tmp_path / "out.xlsx"
        save_excel(_SAMPLE_ENTRIES, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 1).value == "第10104号"
        assert ws.cell(4, 3).value == "連先"   # 3件目 category


# ---------------------------------------------------------------------------
# load_excel_config
# ---------------------------------------------------------------------------


class TestLoadExcelConfig:
    def test_returns_dict(self):
        yaml_path = Path(__file__).parents[1] / "config" / "clients.yaml"
        cfg = load_excel_config(yaml_path)
        assert isinstance(cfg, dict)

    def test_sheet_name(self):
        yaml_path = Path(__file__).parents[1] / "config" / "clients.yaml"
        cfg = load_excel_config(yaml_path)
        assert cfg["sheet_name"] == "相続案件リスト"

    def test_column_order(self):
        yaml_path = Path(__file__).parents[1] / "config" / "clients.yaml"
        cfg = load_excel_config(yaml_path)
        assert cfg["column_order"] == ["受付番号", "受付日", "区分", "目的", "種別", "所在", "備考"]

    def test_column_order_is_list(self):
        yaml_path = Path(__file__).parents[1] / "config" / "clients.yaml"
        cfg = load_excel_config(yaml_path)
        assert isinstance(cfg["column_order"], list)
